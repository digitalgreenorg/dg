import os

from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from videos.models import *
from activities.models import *
from people.models import *
from geographies.models import *

class Command(BaseCommand):

    def handle(self, *args, **options):
        pm_file = '/Users/nikhilverma/workspace/DG/dg/Information_for_COCO_ID.xlsx'
        # path = os.path.abspath(pm_file)
        partner_obj = Partner.objects.get(partner_name="VARRAT")
        partner_id = partner_obj.id
        wb = load_workbook(pm_file)
        # This
        sheet_names = ['Arm-1', 'Arm-2', 'Arm-3']
        # sheet_names = ['UP', 'Ethiopia']

        data_list = []
        for sheet_iterable in sheet_names:
            ws = wb.get_sheet_by_name(sheet_iterable)
            print "wsssss", ws.title
            
            row_num = ws.max_row
            for row_index in range(4,row_num):
                try:
                    data_list.append([{
                                       'mediator_name': ws[row_index][1].value,
                                       'district_name': ws[row_index][2].value,
                                       'block_name': ws[row_index][3].value,
                                       'village_name': ws[row_index][4].value}])
                except Exception as e:
                    pass
                
        for item in data_list:
            item = item[0]
            if item.get('mediator_name') is not None:
                try:
                    district_obj = District.objects.get(district_name=item.get('district_name'))
                    block_obj, created = \
                        Block.objects.get_or_create(district_id=district_obj.id,
                                                    block_name=item.get('block_name'))
                    print "block", created

                    village_obj, created = \
                        Village.objects.get_or_create(village_name=item.get('village_name'),
                                                      block_id=block_obj.id)
                    print "village_obj", created

                    try:
                        mediator_obj, created = \
                            Animator.objects.get_or_create(partner_id=partner_id,
                                                           name=item.get('mediator_name'),
                                                           district_id=district_obj.id)
                    except:
                        mediator_obj = Animator.objects.filter(partner_id=partner_id,
                                                               name=item.get('mediator_name'),
                                                               district_id=district_obj.id)
                        if len(mediator_obj):
                            mediator_obj = mediator_obj.latest('id')

                    print "mediator_obj", created
                    obj, created = AnimatorAssignedVillage.objects.get_or_create(village_id=village_obj.id, animator=mediator_obj)
                    print created, "CREATED"
                except Exception as e:
                    print e