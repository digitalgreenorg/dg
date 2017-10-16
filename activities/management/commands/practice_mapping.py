import os

from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import load_workbook
from videos.models import *

class Command(BaseCommand):

    def handle(self, *args, **options):
        pm_file = '/Users/nikhilverma/workspace/DG/dg/Practice_map2.xlsx'
        # path = os.path.abspath(pm_file)
        wb = load_workbook(pm_file)
        # This
        sheet_names = ['Bihar', 'Rajasthan']
        # sheet_names = ['UP', 'Ethiopia']

        video_data_list = []
        for sheet_iterable in sheet_names:
            ws = wb.get_sheet_by_name(sheet_iterable)
            print "wsssss", ws.title
            
            row_num = ws.max_row
            for row_index in range(3,row_num):
                # print row_index, "------>", row_num, ws[row_index][1].value
                # print ws[row_index][1].value, ws[row_index][6].value, ws[row_index][7].value, ws[row_index][8].value
                # if ws[row_index][1].value is not None and ws[row_index][6].value is not None and ws[row_index][8].value is not None: 
                try:
                    video_data_list.append([{'id': int(ws[row_index][1].value),
                                             'category_name': ws[row_index][6].value,
                                             'sub_category_name': ws[row_index][7].value,
                                             'videopractice_name': ws[row_index][8].value}])
                except Exception as e:
                    print e
            for item in video_data_list:
                try:
                    video_obj = Video.objects.get(id=int(item[0].get('id')))
                    if not video_obj.category_id and item[0].get('category_name'):
                        category_obj = Category.objects.get(category_name=item[0].get('category_name'))
                        video_obj.category_id=category_obj.id
                        video_obj.save()
                        print video_obj.category_id
                    if not video_obj.subcategory_id and item[0].get('sub_category_name'):
                        print video_obj.category_id, item[0].get('category_name'), item[0].get('sub_category_name'), item[0].get('videopractice_name')
                        subcategory_obj = SubCategory.objects.filter(subcategory_name=item[0].get('sub_category_name'))
                        if len(subcategory_obj):
                            video_obj.subcategory_id = subcategory_obj.latest('id')
                            video_obj.save()
                        else:
                            category_obj = Category.objects.get(category_name=item[0].get('category_name'))
                            subcategory_obj, created = SubCategory.objects.get_or_create(subcategory_name=item[0].get('sub_category_name'),
                                                                                         category_id=category_obj.id)
                            print created
                            video_obj.subcategory_id = subcategory_obj.id
                            video_obj.save()
                        print video_obj.subcategory_id
                    if video_obj.videopractice.count() < 1 and item[0].get('videopractice_name'):
                        # subcategory_obj = SubCategory.objects.flter(subcategory_name=item[0].get('sub_category_name'))
                        # if len(subcategory_obj):
                        #     subcategory_obj = subcategory_obj.latest('id')
                        videopractice_obj = VideoPractice.objects.filter(videopractice_name=item[0].get('videopractice_name'))
                                                                         # subcategory__subcategory_name=item[0].get('sub_category_name'))
                        for item in videopractice_obj:
                            video_obj.videopractice.add(item)
                            video_obj.save()
                            print video_obj.videopractice        
                except Exception as e:
                    print e
